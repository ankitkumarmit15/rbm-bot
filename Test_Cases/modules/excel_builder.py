"""
Excel Builder module: Create and format Excel workbooks

Supports two output modes:
  1. build_excel_workbook()         — single Test Cases sheet (5-column template format)
  2. build_full_strategy_workbook() — 7-tab full BOT Test Strategy workbook
"""

import io
import datetime
import openpyxl
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .config import TAB_SCHEMAS

# ─────────────────────────────────────────────────────────────
# Template colour palette (extracted from Wi-Fi Auth Fee template)
# ─────────────────────────────────────────────────────────────
_C_NAVY        = "1F3864"   # dark navy — title text, label text
_C_HEADER_BG   = "9CC2E5"   # light blue — sheet header rows
_C_TITLE_BG    = "1F3864"   # navy fill — cover page title banner
_C_LABEL_BG    = "DEEAF6"   # pale blue — label/key cells
_C_ALT_ROW     = "F2F2F2"   # light grey — alternating data rows
_C_SECTION_BG  = "BDD6EE"   # medium blue — section header rows
_C_WHITE       = "FFFFFF"
_C_ACCENT      = "4A86E8"   # bright blue — tab headings / accents

_FONT_NAME = "Calibri"

# ─────────────────────────────────────────────────────────────
# Internal → template column mapping
# (AI generates 11 internal fields; Excel shows 5 template cols)
# ─────────────────────────────────────────────────────────────
_TEMPLATE_COLS = ["TC ID", "TC Title / Description", "Preconditions", "Execution Steps", "Expected Result"]

def _map_to_template(rows: list) -> list:
    """Convert internal 11-field test case rows to 5-column template format."""
    result = []
    for row in rows:
        name = str(row.get("Test Case Name", "") or row.get("TC Title / Description", "")).strip()
        desc = str(row.get("Description", "")).strip()
        combined = name
        if desc and desc.lower() not in (name.lower(), "n/a", "tbd", "none", ""):
            combined = name + "\n\n" + desc if name else desc
        result.append({
            "TC ID":                  str(row.get("Test Case ID", "") or row.get("TC ID", "")).strip(),
            "TC Title / Description": combined,
            "Preconditions":          str(row.get("Precondition", "") or row.get("Preconditions", "")).strip(),
            "Execution Steps":        str(row.get("Test Step Description", "") or row.get("Execution Steps", "")).strip(),
            "Expected Result":        str(row.get("Test Step Expected Result", "") or row.get("Expected Result", "")).strip(),
        })
    return result


# ─────────────────────────────────────────────────────────────
# Shared style helpers
# ─────────────────────────────────────────────────────────────
def _make_styles():
    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "header_font":  Font(name=_FONT_NAME, size=11, bold=True, color=_C_NAVY),
        "title_font":   Font(name=_FONT_NAME, size=14, bold=True, color=_C_WHITE),
        "label_font":   Font(name=_FONT_NAME, size=11, bold=True, color=_C_NAVY),
        "data_font":    Font(name=_FONT_NAME, size=10, color="000000"),
        "small_font":   Font(name=_FONT_NAME, size=10, bold=False, color=_C_NAVY),
        "header_fill":  PatternFill("solid", fgColor=_C_HEADER_BG),
        "title_fill":   PatternFill("solid", fgColor=_C_TITLE_BG),
        "label_fill":   PatternFill("solid", fgColor=_C_LABEL_BG),
        "section_fill": PatternFill("solid", fgColor=_C_SECTION_BG),
        "alt_fill":     PatternFill("solid", fgColor=_C_ALT_ROW),
        "white_fill":   PatternFill("solid", fgColor=_C_WHITE),
        "align_center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "align_left":   Alignment(horizontal="left",   vertical="top",    wrap_text=True),
        "align_left_m": Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "border":       border,
    }


def _sanitize(value) -> str:
    if value is None:
        return ""
    s = str(value)
    s = re.sub(r"\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])", "", s)
    s = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]", "", s)
    return s


_BLANK_VALUES = {"", "none", "n/a", "tbd", "not available", "blank", "null"}


# ─────────────────────────────────────────────────────────────
# Single-sheet workbook  (replaces old build_excel_workbook)
# ─────────────────────────────────────────────────────────────
def build_excel_workbook(
    all_tab_data: dict,
    selected_tabs: list[str] | None = None,
    coverage_data: dict | None = None,
) -> io.BytesIO:
    """
    Single Test Cases sheet in 5-column template format.
    Includes optional Traceability Matrix if coverage_data supplied.
    """
    S = _make_styles()
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    rows = _map_to_template(all_tab_data.get("Test Cases", []))

    ws = wb.create_sheet(title="Test Cases")
    _write_tc_sheet(ws, rows, S)

    if coverage_data and coverage_data.get("mapping"):
        _build_traceability_sheet(
            wb, all_tab_data.get("Test Cases", []), coverage_data, S
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_tc_sheet(ws, rows: list, S: dict):
    """Write header + data rows to a Test Cases worksheet."""
    # Header row
    for col_idx, hdr in enumerate(_TEMPLATE_COLS, 1):
        c = ws.cell(row=1, column=col_idx, value=hdr)
        c.font    = S["header_font"]
        c.fill    = S["header_fill"]
        c.alignment = S["align_center"]
        c.border  = S["border"]

    # Data rows
    for r_idx, row in enumerate(rows, 2):
        use_alt = (r_idx % 2 == 0)
        for c_idx, col in enumerate(_TEMPLATE_COLS, 1):
            raw = row.get(col, "")
            val = _sanitize(raw)
            if val.lower() in _BLANK_VALUES:
                val = ""
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font      = S["data_font"]
            c.alignment = S["align_left"]
            c.border    = S["border"]
            c.fill      = S["alt_fill"] if use_alt else S["white_fill"]

    # Column widths
    col_widths = [12, 45, 30, 50, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────
# Full 7-tab Test Strategy workbook
# ─────────────────────────────────────────────────────────────
def build_full_strategy_workbook(
    test_cases: list,
    strategy_content: dict,
    coverage_data: dict | None = None,
) -> io.BytesIO:
    """
    Build a 7-tab full BOT Test Strategy workbook matching the project template.

    Tabs (in order):
      1. Cover Page         — AI fills project metadata
      2. Project Overview   — AI fills summary + activities
      3. BOT Timelines      — Headers only (manual fill)
      4. Test Scenarios     — AI fills numbered scenarios
      5. Test Data          — Headers only (manual fill)
      6. Test Cases         — AI fills 5-column TC sheet
      7. Samples            — GL accounts reference structure

    strategy_content keys: project_name, jira_refs, feature_summary,
                           bot_activities, test_scenarios (list of dicts)
    """
    S = _make_styles()
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    sc = strategy_content or {}

    _build_cover_page(wb, sc, S)
    _build_project_overview(wb, sc, S)
    _build_bot_timelines(wb, S)
    _build_test_scenarios(wb, sc, S)
    _build_test_data(wb, S)
    _build_tc_tab(wb, test_cases, S)
    _build_samples(wb, S)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
# Tab builders
# ─────────────────────────────────────────────────────────────

def _build_cover_page(wb, sc: dict, S: dict):
    ws = wb.create_sheet("Cover Page")
    project_name = sc.get("project_name", "BOT Test Documentation")
    jira_refs    = sc.get("jira_refs", "")

    # ── Title banner (row 2, merged A2:Z6) ──
    title_text = f"BOT TEST DOCUMENTATION\n{project_name}"
    if jira_refs:
        title_text += f"\n{jira_refs}"

    ws.merge_cells("A2:Z6")
    c = ws["A2"]
    c.value     = title_text
    c.font      = Font(name=_FONT_NAME, size=16, bold=True, color=_C_WHITE)
    c.fill      = S["title_fill"]
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 60

    # ── Metadata table (rows 7-14) ──
    today = datetime.date.today().strftime("%d-%b-%Y")
    meta_rows = [
        ("Project / Feature",    project_name),
        ("CAP / JIRA Reference", jira_refs or "—"),
        ("Document Type",        "BOT Test Documentation (Test Strategy + Test Plan + Test Cases)"),
        ("Prepared By",          "WQE BOT"),
        ("Reviewed By",          "TBD"),
        ("Date",                 today),
        ("Version",              "1.0"),
        ("Status",               "DRAFT"),
    ]
    _write_label_value_table(ws, start_row=7, rows=meta_rows, S=S)

    # ── ToC section header (row 17) ──
    _section_header(ws, row=17, col=1, text="Document Contents", S=S, merge_end_col=5)

    # ── ToC rows (rows 18-24) ──
    toc_rows = [
        ("Sheet",                 "Description"),
        ("1. Project Overview",   "Overall Project, BOT Scope and Deliverables"),
        ("2. BOT Timelines",      "Sprint plans and development/BOT timeline"),
        ("3. Test Scenarios",     "End-to-End functional test scenarios"),
        ("4. Test Data",          "Provisioning scenario combinations for bill flow"),
        ("5. Test Cases",         "Detailed test cases with steps and expected results"),
        ("6. Samples",            "Reference data: GL accounts, invoice mock-ups"),
    ]
    ws.merge_cells("A18:B18")
    ws.merge_cells("C18:E18")
    for r_offset, (label, desc) in enumerate(toc_rows):
        r = 18 + r_offset
        c_label = ws.cell(row=r, column=1, value=label)
        c_desc  = ws.cell(row=r, column=3, value=desc)
        if r_offset == 0:
            c_label.font  = S["label_font"]
            c_label.fill  = S["section_fill"]
            c_desc.font   = S["label_font"]
            c_desc.fill   = S["section_fill"]
        else:
            c_label.font  = S["data_font"]
            c_label.fill  = S["label_fill"] if r_offset % 2 == 1 else S["white_fill"]
            c_desc.font   = S["data_font"]
            c_desc.fill   = S["label_fill"] if r_offset % 2 == 1 else S["white_fill"]
        for col in [1, 3]:
            ws.cell(row=r, column=col).alignment = S["align_left_m"]
            ws.cell(row=r, column=col).border    = S["border"]

    # ── Version history header (row 27) ──
    _section_header(ws, row=27, col=1, text="Version History", S=S, merge_end_col=5)

    # ── Version history table (row 28 header + rows 29-30 data) ──
    vh_headers = ["Version", "Date", "Description", "Prepared By"]
    for ci, h in enumerate(vh_headers, 1):
        c = ws.cell(row=28, column=ci, value=h)
        c.font  = S["label_font"]
        c.fill  = S["section_fill"]
        c.alignment = S["align_center"]
        c.border = S["border"]

    vh_data = [("1.0", today, "Initial Documentation", "WQE BOT")]
    for r_offset, row_vals in enumerate(vh_data):
        r = 29 + r_offset
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.font      = S["data_font"]
            c.fill      = S["alt_fill"] if r_offset % 2 == 0 else S["white_fill"]
            c.alignment = S["align_left_m"]
            c.border    = S["border"]

    # ── Column widths ──
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 20


def _build_project_overview(wb, sc: dict, S: dict):
    ws = wb.create_sheet("Project Overview")
    project_name    = sc.get("project_name", "—")
    feature_summary = sc.get("feature_summary", "—")
    bot_activities  = sc.get("bot_activities", "—")

    # ── Key info rows ──
    info_rows = [
        ("Project Name",       project_name),
        ("Solution Summary",   feature_summary),
        ("BOT Testing Activities", bot_activities),
    ]
    _write_label_value_table(ws, start_row=2, rows=info_rows, S=S, value_col=2, value_merge_end=5)

    # ── Partner table header (row 7) ──
    _section_header(ws, row=7, col=1, text="Partners in Scope", S=S, merge_end_col=5)

    partner_headers = ["Scenarios", "Partner Name", "Partner Type", "Config. Type", "BCC/ACC"]
    for ci, h in enumerate(partner_headers, 1):
        c = ws.cell(row=8, column=ci, value=h)
        c.font      = S["label_font"]
        c.fill      = S["header_fill"]
        c.alignment = S["align_center"]
        c.border    = S["border"]

    # Empty partner rows (3 placeholder rows)
    for r in range(9, 12):
        for ci in range(1, 6):
            c = ws.cell(row=r, column=ci, value="")
            c.fill   = S["alt_fill"] if r % 2 == 0 else S["white_fill"]
            c.border = S["border"]

    # ── BOT Deliverables section (row 13) ──
    _section_header(ws, row=13, col=1, text="BOT Deliverables", S=S, merge_end_col=5)

    deliv_headers = ["Operations Reports", "Accounting Reports", "Tax Reports", "Accruals Reports"]
    for ci, h in enumerate(deliv_headers, 1):
        c = ws.cell(row=14, column=ci, value=h)
        c.font      = S["label_font"]
        c.fill      = S["header_fill"]
        c.alignment = S["align_center"]
        c.border    = S["border"]

    # Empty deliverable rows (5 placeholder rows)
    for r in range(15, 20):
        for ci in range(1, 5):
            c = ws.cell(row=r, column=ci, value="")
            c.fill   = S["alt_fill"] if r % 2 == 0 else S["white_fill"]
            c.border = S["border"]

    # ── Column widths ──
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.row_dimensions[3].height = 90   # Solution Summary — tall row
    ws.row_dimensions[4].height = 60   # BOT Activities


def _build_bot_timelines(wb, S: dict):
    ws = wb.create_sheet("BOT Timelines")
    headers = [
        "TMS", "Issue key", "Summary", "Status", "REL", "WQE CD",
        "Drop", "", "BOT Sprint", "Sprint Start", "Sprint End",
        "Release", "BOT U/S", "BOT U/S Desc",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        if h:
            c.font      = S["header_font"]
            c.fill      = S["header_fill"]
            c.alignment = S["align_center"]
            c.border    = S["border"]

    # Empty data rows (10 placeholder rows)
    for r in range(2, 12):
        for ci in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=ci, value="")
            c.fill   = S["alt_fill"] if r % 2 == 0 else S["white_fill"]
            c.border = S["border"]

    col_widths = [18, 14, 55, 8, 8, 10, 14, 4, 12, 13, 13, 10, 12, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _build_test_scenarios(wb, sc: dict, S: dict):
    ws = wb.create_sheet("Test Scenarios")

    # Header row
    headers = ["SL. No.", "", "Scenario Name", "Scenario Expected Result"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        if h:
            c.font      = S["header_font"]
            c.fill      = S["header_fill"]
            c.alignment = S["align_center"]
            c.border    = S["border"]

    scenarios = sc.get("test_scenarios", [])
    r = 2
    for scen in scenarios:
        sl_no = str(scen.get("sl_no", "")).strip()
        name  = _sanitize(scen.get("scenario_name", ""))
        exp   = _sanitize(scen.get("expected_result", ""))

        # Determine if top-level (integer) or sub-scenario (has decimal)
        is_top = "." not in sl_no
        use_alt = (r % 2 == 0)

        if is_top:
            a_val, b_val = sl_no, ""
            fill = S["label_fill"] if is_top else (S["alt_fill"] if use_alt else S["white_fill"])
            fnt  = S["label_font"] if is_top else S["data_font"]
        else:
            a_val, b_val = "", sl_no
            fill = S["alt_fill"] if use_alt else S["white_fill"]
            fnt  = S["data_font"]

        cells_data = [(1, a_val), (2, b_val), (3, name), (4, exp)]
        for ci, val in cells_data:
            c = ws.cell(row=r, column=ci, value=val)
            c.font      = fnt
            c.fill      = fill
            c.alignment = S["align_left"]
            c.border    = S["border"]
        ws.row_dimensions[r].height = max(15, min(60, len(name) // 3 + 15))
        r += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A2"


def _build_test_data(wb, S: dict):
    ws = wb.create_sheet("Test Data")
    headers = [
        "SL.No", "Action", "Plan", "Subscriber ID", "MSISDN",
        "On Day ( of Bill Cycle)", "Status",
        "Wifi Charge Applied (Yes/No)", "Charge Count", "Final Plan for Billing",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = S["header_font"]
        c.fill      = S["header_fill"]
        c.alignment = S["align_center"]
        c.border    = S["border"]

    # Empty data rows (10 placeholder rows)
    for r in range(2, 12):
        for ci in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=ci, value="")
            c.fill   = S["alt_fill"] if r % 2 == 0 else S["white_fill"]
            c.border = S["border"]

    col_widths = [8, 14, 16, 16, 14, 20, 12, 24, 14, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _build_tc_tab(wb, test_cases: list, S: dict):
    ws = wb.create_sheet("Test Cases")
    rows = _map_to_template(test_cases)
    _write_tc_sheet(ws, rows, S)


def _build_samples(wb, S: dict):
    ws = wb.create_sheet("Samples")

    # Invoice mock-up header
    _section_header(ws, row=1, col=1, text="Invoice Mock-up", S=S, merge_end_col=8)

    # Blank invoice rows
    for r in range(2, 12):
        for ci in range(1, 9):
            c = ws.cell(row=r, column=ci, value="")
            c.fill   = S["white_fill"]
            c.border = S["border"]

    # GL Accounts section
    _section_header(ws, row=13, col=1, text="GL Accounts", S=S, merge_end_col=8)

    gl_headers = ["Line Item Text", "GL Account", "Category", "Reporting Ind",
                  "", "Read Only Flag", "Account Type", "Line Text"]
    for ci, h in enumerate(gl_headers, 1):
        c = ws.cell(row=14, column=ci, value=h)
        if h:
            c.font      = S["label_font"]
            c.fill      = S["header_fill"]
            c.alignment = S["align_center"]
            c.border    = S["border"]

    # Empty GL data rows
    for r in range(15, 22):
        for ci in range(1, 9):
            c = ws.cell(row=r, column=ci, value="")
            c.fill   = S["alt_fill"] if r % 2 == 0 else S["white_fill"]
            c.border = S["border"]

    col_widths = [35, 14, 12, 14, 4, 14, 16, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─────────────────────────────────────────────────────────────
# Shared layout helpers
# ─────────────────────────────────────────────────────────────

def _section_header(ws, row: int, col: int, text: str, S: dict, merge_end_col: int = 4):
    """Write a full-width section header with navy fill + white bold text."""
    end_col = get_column_letter(merge_end_col)
    start_col = get_column_letter(col)
    ws.merge_cells(f"{start_col}{row}:{end_col}{row}")
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name=_FONT_NAME, size=12, bold=True, color=_C_WHITE)
    c.fill      = PatternFill("solid", fgColor=_C_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def _write_label_value_table(
    ws, start_row: int, rows: list,
    S: dict, label_col: int = 1, value_col: int = 3,
    value_merge_end: int = 5,
):
    """Write rows of (label, value) pairs with label-fill on col A and data in col C."""
    end_col = get_column_letter(value_merge_end)
    val_start = get_column_letter(value_col)
    for r_offset, (label, value) in enumerate(rows):
        r = start_row + r_offset
        use_alt = (r_offset % 2 == 1)

        # Merge value cells
        if value_col < value_merge_end:
            ws.merge_cells(f"{val_start}{r}:{end_col}{r}")

        c_label = ws.cell(row=r, column=label_col, value=label)
        c_label.font      = S["label_font"]
        c_label.fill      = S["label_fill"]
        c_label.alignment = S["align_left_m"]
        c_label.border    = S["border"]

        c_val = ws.cell(row=r, column=value_col, value=_sanitize(value))
        c_val.font      = S["data_font"]
        c_val.fill      = S["alt_fill"] if use_alt else S["white_fill"]
        c_val.alignment = S["align_left"]
        c_val.border    = S["border"]


# ─────────────────────────────────────────────────────────────
# Traceability Matrix (unchanged from original)
# ─────────────────────────────────────────────────────────────

def _build_traceability_sheet(wb, test_cases, coverage_data, S):
    from openpyxl.styles import PatternFill as _PF

    ws = wb.create_sheet("Traceability Matrix")
    mapping = coverage_data.get("mapping", {})
    parsed_reqs = coverage_data.get("parsed_requirements", [])

    tc_ids   = [tc.get("Test Case ID", f"TC-{i+1:03d}") for i, tc in enumerate(test_cases)]
    tc_names = [tc.get("Test Case Name", "") for tc in test_cases]

    req_lookup = {r.get("req_id", ""): r for r in parsed_reqs}
    for rid in mapping:
        if rid not in req_lookup:
            req_lookup[rid] = {"req_id": rid, "action": "", "priority": ""}

    header = ["Requirement ID", "Description", "Priority"] + tc_ids
    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font      = S["header_font"]
        c.fill      = S["header_fill"]
        c.alignment = S["align_center"]
        c.border    = S["border"]

    green_fill = _PF(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    check_font = Font(name=_FONT_NAME, size=10, bold=True, color="276221")

    for req_id, covered_tc_names in sorted(mapping.items()):
        covered_set = set(covered_tc_names)
        meta = req_lookup.get(req_id, {})
        action   = str(meta.get("action", ""))[:80]
        priority = str(meta.get("priority", ""))

        row_vals = [req_id, action, priority] + [
            "✓" if name in covered_set else "" for name in tc_names
        ]
        ws.append(row_vals)
        row_num = ws.max_row

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font      = S["data_font"]
            cell.alignment = S["align_left_m"]
            cell.border    = S["border"]
            if val == "✓":
                cell.fill      = green_fill
                cell.font      = check_font
                cell.alignment = S["align_center"]

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    for col_idx in range(4, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 9
    ws.freeze_panes = "A2"
